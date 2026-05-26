"""
Layer 3 — FastAPI Backend
All six endpoints for ChurnShield.
"""

import os
import io
import sys
import math
from dotenv import load_dotenv
load_dotenv()   # loads .env from the project root before anything else reads os.environ
import numpy as np
import pandas as pd
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def _clean_for_json(obj):
    """Recursively replace NaN/Inf with None so JSON serialization works."""
    if isinstance(obj, dict):
        return {k: _clean_for_json(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [_clean_for_json(v) for v in obj]
    elif isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
        return 0.0
    return obj


# ──────────────────── EXCEL FORMATTING HELPERS ────────────────────

# Cell fill colours
_FILL_HEADER = PatternFill("solid", fgColor="1F3864")   # dark navy
_FILL_HIGH   = PatternFill("solid", fgColor="FFCCCC")   # light red
_FILL_MEDIUM = PatternFill("solid", fgColor="FFECC2")   # light orange
_FILL_LOW    = PatternFill("solid", fgColor="D5F5D5")   # light green
_FILL_ALT    = PatternFill("solid", fgColor="F2F2F2")   # alternating row grey

_FONT_HEADER = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
_FONT_BOLD   = Font(bold=True, name="Calibri", size=10)
_FONT_BODY   = Font(name="Calibri", size=10)

_RISK_FILL = {"High": _FILL_HIGH, "Medium": _FILL_MEDIUM, "Low": _FILL_LOW}


def _auto_fit_columns(ws, min_width: int = 10, max_width: int = 55):
    """Adjust column widths to the longest value in each column."""
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = min(max(max_len + 3, min_width), max_width)


def _style_predictions_sheet(ws):
    """
    Format the Predictions worksheet:
      • Dark navy bold header row (frozen)
      • Risk_Level cells colour-coded: High=red, Medium=orange, Low=green
      • Alternating row shading on all other columns
      • Right-aligned probability-percentage column
      • Auto-fit column widths
    """
    # Map column names → column indices
    header_map = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
    risk_col     = header_map.get("Risk_Level")
    prob_pct_col = header_map.get("Churn_Probability_%")

    # -- Header row --
    for cell in ws[1]:
        cell.font      = _FONT_HEADER
        cell.fill      = _FILL_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # -- Data rows --
    for row_idx in range(2, ws.max_row + 1):
        is_alt = (row_idx % 2 == 0)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.font      = _FONT_BODY
            cell.alignment = Alignment(vertical="center")
            if is_alt and col_idx != risk_col:
                cell.fill = _FILL_ALT

        # Colour Risk_Level cell by tier
        if risk_col:
            rc   = ws.cell(row_idx, risk_col)
            fill = _RISK_FILL.get(str(rc.value) if rc.value else "")
            if fill:
                rc.fill      = fill
                rc.font      = _FONT_BOLD
                rc.alignment = Alignment(horizontal="center", vertical="center")

        # Right-align probability percentage
        if prob_pct_col:
            ws.cell(row_idx, prob_pct_col).alignment = Alignment(
                horizontal="right", vertical="center"
            )

    _auto_fit_columns(ws)


def _write_summary_sheet(writer: pd.ExcelWriter, result_df: pd.DataFrame):
    """Add a Summary sheet with risk-tier counts, percentages, and overall stats."""
    total = len(result_df)
    risk_counts = (
        result_df["Risk_Level"]
        .astype(str)
        .value_counts()
        .reindex(["High", "Medium", "Low"], fill_value=0)
    )

    tier_df = pd.DataFrame({
        "Risk Tier":    risk_counts.index.tolist(),
        "Customers":    risk_counts.values.tolist(),
        "Share (%)": [round(n / total * 100, 1) for n in risk_counts.values],
    })

    avg_prob = round(float(result_df["Churn_Probability"].mean()) * 100, 1)
    stats_df = pd.DataFrame({
        "Metric": ["Total Customers Analysed", "Mean Churn Probability (%)"],
        "Value":  [total, avg_prob],
    })

    start_row_stats = len(tier_df) + 3          # blank row gap
    tier_df.to_excel(writer,  sheet_name="Summary", index=False, startrow=0)
    stats_df.to_excel(writer, sheet_name="Summary", index=False, startrow=start_row_stats)

    ws = writer.sheets["Summary"]

    # Style both header rows
    for header_row in (1, start_row_stats + 1):
        for cell in ws[header_row]:
            cell.font      = _FONT_HEADER
            cell.fill      = _FILL_HEADER
            cell.alignment = Alignment(horizontal="center")

    # Colour tier cells in the breakdown table
    for row_idx in range(2, len(tier_df) + 2):
        cell = ws.cell(row_idx, 1)
        fill = _RISK_FILL.get(str(cell.value) if cell.value else "")
        if fill:
            cell.fill = fill
            cell.font = _FONT_BOLD

    _auto_fit_columns(ws)


# Ensure project root is on path so we can import sibling packages
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from src.predict import predict as xgb_predict
from backend.suggestions import get_suggestion
from backend.revenue import (
    calculate_revenue_impact,
    calculate_customer_revenue_risk,
    calculate_roi,
)
from backend.message_generator import generate_message, generate_and_send
from backend.analytics import build_analytics_response, build_trends_response

# ──────────────────── DATA PATH + CACHED DATAFRAME ────────────────────
DATA_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "ecommerce_churn_enhanced.csv")

# Load the dataset once at startup and cache it in memory.
# All analytics endpoints read from this shared, immutable DataFrame.
# Re-reads happen only if the process is restarted (acceptable for a CSV-backed app).
try:
    _df_cache: pd.DataFrame = pd.read_csv(DATA_PATH)
except FileNotFoundError:
    _df_cache = pd.DataFrame()  # empty sentinel; endpoints will raise 503

# ──────────────────── APP ────────────────────
app = FastAPI(
    title="ChurnShield API",
    description="Customer Churn Prediction & Retention Intelligence",
    version="1.0.0",
)

ALLOWED_ORIGINS = [
    "http://localhost:8080",   # Vite dev server
    "http://localhost:4173",   # Vite preview (vite preview)
    "http://127.0.0.1:8080",
    "http://127.0.0.1:4173",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization", "Accept"],
)


# ──────────────────── SCHEMAS ────────────────────

class CustomerInput(BaseModel):
    Tenure: Optional[float] = 12
    PreferredLoginDevice: Optional[str] = "Mobile Phone"
    CityTier: Optional[int] = 1
    WarehouseToHome: Optional[float] = 15
    PreferredPaymentMode: Optional[str] = "Debit Card"
    Gender: Optional[str] = "Male"
    HourSpendOnApp: Optional[float] = 3
    NumberOfDeviceRegistered: Optional[int] = 3
    PreferedOrderCat: Optional[str] = "Laptop & Accessory"
    SatisfactionScore: Optional[int] = 3
    MaritalStatus: Optional[str] = "Single"
    NumberOfAddress: Optional[int] = 2
    Complain: Optional[int] = 0
    OrderAmountHikeFromlastYear: Optional[float] = 15
    CouponUsed: Optional[float] = 1
    OrderCount: Optional[float] = 2
    DaySinceLastOrder: Optional[float] = 5
    CashbackAmount: Optional[float] = 150
    TotalSpend: Optional[float] = 0.0
    AvgOrderValue: Optional[float] = 0.0
    ReturnRate: Optional[float] = 0.0
    CustomerAge: Optional[int] = 30
    LastLoginDaysAgo: Optional[int] = 5
    ReviewsGiven: Optional[int] = 0
    WishlistItems: Optional[int] = 0
    SubscriptionPlan: Optional[str] = "Free"
    ReferralsMade: Optional[int] = 0
    SupportTicketCount: Optional[int] = 0


class RevenueInput(BaseModel):
    at_risk_customers: int
    avg_order_value: float
    coupon_amount: float
    retention_rate: float  # percentage, e.g. 30
    orders_per_year: float = 1.0  # annualised orders per retained customer


# ── /metrics/revenue-impact schemas ──────────────────────────────────────────

class CustomerRevenueRecord(BaseModel):
    """
    One customer's data needed for revenue-risk calculation.
    """
    customer_id:        Optional[str]   = None
    churn_probability:  float           # 0.0 – 1.0  (from /predict or /predict/bulk)
    revenue_value:      float           # per-order revenue (£ / $ / ₹)
    orders_per_year:    float           = 1.0
    risk_level:         Optional[str]   = None  # pre-classified; computed if omitted


class RevenueImpactInput(BaseModel):
    """
    Full payload for POST /metrics/revenue-impact.
    Provide a list of customers (with individual churn probabilities and revenue
    values) plus campaign parameters to get a complete ROI analysis.
    """
    customers:                  list[CustomerRevenueRecord]
    campaign_cost_per_customer: float   # cost of reaching one customer (coupon / call / email)
    retention_rate:             float   # expected % of targeted customers retained, e.g. 30


class SuggestInput(BaseModel):
    DaySinceLastOrder: Optional[float] = 5
    SatisfactionScore: Optional[int] = 3
    Complain: Optional[int] = 0
    CashbackAmount: Optional[float] = 150
    Tenure: Optional[float] = 12


class MessageInput(BaseModel):
    customer_segment: str
    suggestion: str
    tone: Optional[str] = "warm, concise"
    channel: Optional[str] = "WhatsApp"
    churn_signals: Optional[list] = None   # ranked list from /suggest


class MessageSendInput(BaseModel):
    customer_segment: str
    suggestion: str
    to_number: str           # E.164, e.g. '+919876543210'
    tone: Optional[str] = "warm, concise"
    channel: Optional[str] = "WhatsApp"
    churn_signals: Optional[list] = None


# ──────────────────── ENDPOINTS ────────────────────

# ---------- 1. POST /predict — Single Customer Prediction ----------
@app.post("/predict")
def predict_customer(customer: CustomerInput):
    """Module 1: Predict churn for a single customer."""
    data = customer.model_dump()
    result = xgb_predict(data)
    return result


# ---------- 2. POST /predict/bulk — Bulk CSV Prediction ----------
@app.post("/predict/bulk")
async def bulk_predict(file: UploadFile = File(...)):
    """
    Upload a CSV of customer records, run batch churn predictions, and
    receive a formatted Excel file with risk scores and recommended actions.

    Input:  multipart/form-data, field name: 'file', CSV format
    Output: .xlsx with two sheets — 'Predictions' and 'Summary'
    """
    filename = file.filename or "upload"

    # ── 1. Validate file type ────────────────────────────────────
    if not filename.lower().endswith(".csv"):
        raise HTTPException(status_code=400, detail="Only .csv files are accepted.")

    # ── 2. Parse CSV ──────────────────────────────────────────────
    contents = await file.read()
    if not contents:
        raise HTTPException(status_code=400, detail="Uploaded file is empty.")

    try:
        df = pd.read_csv(io.BytesIO(contents))
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Could not parse CSV: {exc}")

    if df.empty:
        raise HTTPException(status_code=422, detail="CSV contains no data rows.")

    # ── 3. Batch predictions ──────────────────────────────────────
    result_df = xgb_predict(df)

    # ── 4. Enrich output columns ──────────────────────────────────
    # Rename Suggested_Action → Recommended_Action (clearer for end-users)
    result_df = result_df.rename(columns={"Suggested_Action": "Recommended_Action"})

    # Insert Churn_Probability_% immediately after Risk_Level
    prob_pct = (result_df["Churn_Probability"] * 100).round(1)
    risk_col_pos = result_df.columns.get_loc("Risk_Level") + 1
    result_df.insert(risk_col_pos, "Churn_Probability_%", prob_pct)

    # ── 5. Write formatted Excel into memory ──────────────────────
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        result_df.to_excel(writer, index=False, sheet_name="Predictions")
        _write_summary_sheet(writer, result_df)
        _style_predictions_sheet(writer.sheets["Predictions"])
    output.seek(0)

    stem = os.path.splitext(filename)[0]
    download_name = f"churnshield_{stem}_results.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{download_name}"'},
    )


# Backward-compat alias — redirects old /bulk calls to /predict/bulk
@app.post("/bulk")
async def bulk_predict_alias(file: UploadFile = File(...)):
    """Deprecated alias for /predict/bulk — kept for backward compatibility."""
    return await bulk_predict(file)


# ---------- 3. GET /analytics — Live Analytics Dashboard Data ----------
@app.get("/analytics")
def get_analytics():
    """
    Module 3: Full aggregated analytics payload for the dashboard.

    Returns
    -------
    {
      total_customers         int
      churned_customers       int
      overall_churn_rate      float   (%)
      churn_by_city_tier      {tier: pct}
      churn_by_gender         {gender: pct}
      churn_by_satisfaction   {score: pct}
      churn_by_device         {device: pct}
      churn_by_marital_status {status: pct}
      churn_by_category       {category: pct}
      churn_by_payment_mode   {mode: pct}
      churn_by_tenure         {band: pct}
      avg_days_since_last_order  {churned: float, stayed: float}
      kpi_comparison          {metric: {churned: float, stayed: float}}
    }
    """
    if _df_cache.empty:
        raise HTTPException(status_code=503, detail="Dataset not available.")
    return _clean_for_json(build_analytics_response(_df_cache))


@app.get("/analytics/trends")
def get_analytics_trends():
    """
    Module 3b: Monthly churn time-series using Tenure as the lifecycle axis.

    The dataset has no calendar date column so `Tenure` (months on platform,
    0-60) serves as the time dimension.  At each month T, the churn_rate is
    the proportion of customers with exactly Tenure == T who churned — this
    is the customer lifecycle hazard curve.

    Returns
    -------
    {
      monthly_trend : [
        { month, churned, stayed, total, churn_rate, rolling_rate }, ...
      ]
      rolling_window         int
      peak_churn_month       {month, churn_rate}
      stabilizes_after_month int  — first month where rolling_rate < 20%%
    }
    """
    if _df_cache.empty:
        raise HTTPException(status_code=503, detail="Dataset not available.")
    return _clean_for_json(build_trends_response(_df_cache))


# ---------- 4. POST /revenue — Revenue Impact Calculator ----------
@app.post("/revenue")
def revenue_impact(data: RevenueInput):
    """Module 4: Calculate revenue at risk and campaign ROI."""
    return calculate_revenue_impact(
        at_risk_customers=data.at_risk_customers,
        avg_order_value=data.avg_order_value,
        coupon_amount=data.coupon_amount,
        retention_rate=data.retention_rate,
        orders_per_year=data.orders_per_year,
    )


# ---------- 4b. POST /metrics/revenue-impact — Per-Customer Revenue Risk + ROI ----------
@app.post("/metrics/revenue-impact")
def metrics_revenue_impact(data: RevenueImpactInput):
    """
    Granular revenue-risk analysis with full ROI calculation.

    Flow
    ----
    1. For each customer: at_risk_revenue = churn_probability × revenue_value × orders_per_year
    2. Aggregate totals + tier breakdown (High / Medium / Low)
    3. ROI analysis: campaign_cost, revenue_saved, net_roi, payback_ratio, break-even rate

    Input
    -----
    customers                   list of {customer_id?, churn_probability, revenue_value,
                                         orders_per_year?, risk_level?}
    campaign_cost_per_customer  float  — cost per targeted customer (e.g. coupon value)
    retention_rate              float  — expected retention % from the campaign, e.g. 30

    Output
    ------
    {
      risk_analysis: { total_customers, total_revenue_base, total_revenue_at_risk,
                       weighted_churn_probability_pct, by_risk_tier, top_at_risk_customers },
      roi:           { campaign_cost, customers_retained, revenue_saved, net_roi,
                       roi_percentage, payback_ratio, break_even_retention_rate_pct,
                       is_roi_positive },
      meta:          { retention_rate_input, campaign_cost_per_customer }
    }
    """
    if not data.customers:
        raise HTTPException(status_code=422, detail="customers list must not be empty.")

    # Validate probability values
    bad = [
        i for i, c in enumerate(data.customers)
        if not (0.0 <= c.churn_probability <= 1.0)
    ]
    if bad:
        raise HTTPException(
            status_code=422,
            detail=f"churn_probability must be 0–1. Bad indices: {bad[:5]}",
        )

    # ── Step 1 + 2: per-customer at-risk revenue + aggregation ───
    customer_dicts = [
        {
            "customer_id":       c.customer_id,
            "churn_probability": c.churn_probability,
            "revenue_value":     c.revenue_value,
            "orders_per_year":   c.orders_per_year,
            "risk_level":        c.risk_level,
        }
        for c in data.customers
    ]
    risk_analysis = calculate_customer_revenue_risk(customer_dicts)

    # ── Step 3: ROI ───────────────────────────────────────────────
    roi = calculate_roi(
        total_at_risk_customers    = risk_analysis["total_customers"],
        total_revenue_at_risk      = risk_analysis["total_revenue_at_risk"],
        campaign_cost_per_customer = data.campaign_cost_per_customer,
        retention_rate             = data.retention_rate,
    )

    return _clean_for_json({
        "risk_analysis": risk_analysis,
        "roi":           roi,
        "meta": {
            "retention_rate_input":       data.retention_rate,
            "campaign_cost_per_customer": data.campaign_cost_per_customer,
        },
    })


# ---------- 5. POST /suggest — Smart Retention Suggestion ----------
@app.post("/suggest")
def suggest_retention(customer: SuggestInput):
    """Module 5: Get a rule-based retention suggestion."""
    return get_suggestion(customer.dict())


# ---------- 6. POST /message — AI Message Generator ----------
@app.post("/message")
def generate_retention_message(data: MessageInput):
    """Module 6: Generate a personalised retention message (Ollama → Claude → template)."""
    return generate_message(
        customer_segment = data.customer_segment,
        suggestion       = data.suggestion,
        tone             = data.tone,
        channel          = data.channel or "WhatsApp",
        churn_signals    = data.churn_signals,
    )


# ---------- 6b. POST /message/send — Generate + dispatch via Twilio WhatsApp ----------
@app.post("/message/send")
def send_retention_message(data: MessageSendInput):
    """
    Generate a retention message and immediately send it via Twilio WhatsApp.

    Requires TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN environment variables.

    Input
    -----
    customer_segment  str   e.g. 'inactive', 'complaint'
    suggestion        str   retention offer text
    to_number         str   recipient E.164 phone, e.g. '+919876543210'
    tone              str   (optional) tone instruction
    channel           str   (optional, default 'WhatsApp') label used in LLM prompt
    churn_signals     list  (optional) from POST /suggest response

    Output
    ------
    {
      message  : str   -- generated message text
      source   : str   -- 'ollama' | 'claude' | 'template'
      delivery : {
        success : bool
        sid     : str | null   -- Twilio message SID
        status  : str
        error   : str | null
      }
    }
    """
    return generate_and_send(
        customer_segment = data.customer_segment,
        suggestion       = data.suggestion,
        to_number        = data.to_number,
        tone             = data.tone or "warm, concise",
        channel          = data.channel or "WhatsApp",
        churn_signals    = data.churn_signals,
    )


# ──────────────────── HEALTH CHECK ────────────────────
@app.get("/")
def root():
    return {"status": "ChurnShield API is running", "version": "1.0.0"}


# ──────────────────── RUN ────────────────────
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
